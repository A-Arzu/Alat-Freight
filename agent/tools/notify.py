"""Dispatcher notification: HTML email + XLSX attachment.

Recipients are chosen at runtime from the dashboard (stored in the Firestore
meta document) and fall back to the EMAIL_TO env var. Mail goes out over SMTP
when credentials are configured (SMTP_HOST/SMTP_USER/SMTP_PASS, e.g. a Gmail
app password held in Secret Manager on Cloud Run). The rendered email is always
recorded in the store, so the dashboard can show the deliverable even with no
SMTP configured at all.
"""
import os
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from core.clock import hhmm

PRIORITY_LABEL = {1: "P1 - URGENT", 2: "P2 - LOAD ASAP", 3: "P3 - TODAY"}
PRIORITY_COLOR = {1: "#d64545", 2: "#d69a2e", 3: "#3a7bd5"}

EMAIL_RE = re.compile(r"^[^@\s,;<>]+@[^@\s,;<>]+\.[A-Za-z]{2,}$")
MAX_RECIPIENTS = 5


def parse_recipients(raw: str, limit: int = MAX_RECIPIENTS) -> list[str]:
    """Validate a comma/semicolon separated recipient list.

    Raises ValueError with a message meant to be shown to the user.
    """
    parts = [p.strip() for p in re.split(r"[,;]", raw or "") if p.strip()]
    if not parts:
        raise ValueError("Enter an email address.")
    if len(parts) > limit:
        raise ValueError(f"At most {limit} recipients.")
    for p in parts:
        if len(p) > 254 or not EMAIL_RE.match(p):
            raise ValueError(f"'{p}' is not a valid email address.")
    return parts


def smtp_configured() -> bool:
    return all(os.environ.get(k) for k in ("SMTP_HOST", "SMTP_USER", "SMTP_PASS"))


def resolve_recipients(meta: dict | None, override: str | None = None) -> list[str]:
    """Runtime setting wins, then the stored dashboard setting, then env."""
    for candidate in (override, (meta or {}).get("email_to"), os.environ.get("EMAIL_TO")):
        if candidate and str(candidate).strip():
            try:
                return parse_recipients(str(candidate))
            except ValueError:
                continue
    return []


def build_email_html(plan: dict, meta: dict) -> str:
    rows = []
    for a in sorted(plan["assignments"], key=lambda x: (x["priority"], x["load_start"])):
        chip = (f"<span style='background:{PRIORITY_COLOR[a['priority']]};color:#fff;"
                f"padding:2px 8px;border-radius:10px;font-size:11px'>{PRIORITY_LABEL[a['priority']]}</span>")
        rows.append(
            "<tr>"
            f"<td style='padding:8px;border-bottom:1px solid #e3e8ee'>{chip}</td>"
            f"<td style='padding:8px;border-bottom:1px solid #e3e8ee'><b>{a['shipment_id']}</b></td>"
            f"<td style='padding:8px;border-bottom:1px solid #e3e8ee'>{a['wagon_id']}</td>"
            f"<td style='padding:8px;border-bottom:1px solid #e3e8ee'>{a['team_id']}</td>"
            f"<td style='padding:8px;border-bottom:1px solid #e3e8ee'>{hhmm(a['load_start'])}&ndash;{hhmm(a['load_end'])}</td>"
            f"<td style='padding:8px;border-bottom:1px solid #e3e8ee'>{a['target_ship']}</td>"
            f"<td style='padding:8px;border-bottom:1px solid #e3e8ee'>{a['confidence']}%</td>"
            f"<td style='padding:8px;border-bottom:1px solid #e3e8ee;color:#5b6b77'>{a['reason']}</td>"
            "</tr>")

    holds = "".join(
        f"<li><b>{h['shipment_id']}</b>: {h['action']} &mdash; <i>{h['reason']}</i></li>"
        for h in plan.get("holds", []))

    s = plan.get("summary", {})
    version_note = "" if plan["version"] == 1 else (
        f"<p style='background:#fdf1e7;border-left:4px solid #c85a10;padding:10px 14px'>"
        f"<b>Updated plan v{plan['version']}</b> &mdash; disruption response. "
        f"Changes vs previous plan are listed in the dashboard.</p>")

    return f"""
<div style="font-family:Arial,Helvetica,sans-serif;color:#1b2733;max-width:860px">
  <h2 style="margin:0 0 4px">Daily Dispatch Plan &mdash; {plan['plan_date']}</h2>
  <p style="margin:0 0 12px;color:#5b6b77">Generated {plan['generated_at'].replace('T', ' ')}
     by {plan['planner']} &middot; {meta.get('port_name', 'Port Terminal')}</p>
  {version_note}
  <table style="border-collapse:collapse;width:100%;font-size:13px">
    <tr style="background:#16324a;color:#fff;text-align:left">
      <th style="padding:8px">Priority</th><th style="padding:8px">Shipment</th>
      <th style="padding:8px">Wagon</th><th style="padding:8px">Dock team</th>
      <th style="padding:8px">Load window</th><th style="padding:8px">Ship</th>
      <th style="padding:8px">Conf.</th><th style="padding:8px">Reason</th>
    </tr>
    {''.join(rows)}
  </table>
  {f"<h3 style='margin:16px 0 6px'>On hold</h3><ul>{holds}</ul>" if holds else ""}
  <h3 style="margin:16px 0 6px">Summary</h3>
  <p style="margin:0;color:#334">
    {s.get('planned', 0)} loads scheduled &middot; {s.get('holds', 0)} on hold &middot;
    SLA compliance {s.get('sla_met_pct', '-')}% &middot;
    peak dock utilization {s.get('peak_util_pct', '-')}% &middot;
    planned in {s.get('planning_seconds', '-')}s (manual baseline {meta.get('manual_baseline_min', 45)} min)
  </p>
  <p style="color:#5b6b77;margin-top:14px">Reply to this email or use the dashboard to approve / override.</p>
</div>"""


def build_xlsx(plan: dict, path: str) -> str | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispatch Plan"
    ws.append([f"DISPATCH PLAN {plan['plan_date']} v{plan['version']}"])
    ws["A1"].font = Font(size=13, bold=True)
    headers = ["Priority", "Shipment", "Wagon", "Dock team", "Load start", "Load end",
               "Ship", "Confidence", "Reason"]
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = PatternFill(start_color="16324A", end_color="16324A", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    for a in sorted(plan["assignments"], key=lambda x: (x["priority"], x["load_start"])):
        ws.append([a["priority"], a["shipment_id"], a["wagon_id"], a["team_id"],
                   hhmm(a["load_start"]), hhmm(a["load_end"]), a["target_ship"],
                   f"{a['confidence']}%", a["reason"]])
    for h in plan.get("holds", []):
        ws.append(["HOLD", h["shipment_id"], "-", "-", "-", "-",
                   h.get("rebook_ship") or "-", f"{h['confidence']}%", h["reason"]])
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)
    wb.save(path)
    return path


def _deliver(msg, user: str) -> None:
    """Port 465 = implicit SSL (Gmail default); anything else = STARTTLS."""
    host = os.environ["SMTP_HOST"]
    port = int(os.environ.get("SMTP_PORT", "465"))
    password = os.environ["SMTP_PASS"]
    if port == 465:
        with smtplib.SMTP_SSL(host, port, timeout=30) as smtp:
            smtp.login(user, password)
            smtp.send_message(msg)
    else:
        with smtplib.SMTP(host, port, timeout=30) as smtp:
            smtp.starttls()
            smtp.login(user, password)
            smtp.send_message(msg)


def send_dispatch_email(plan: dict, meta: dict, store, out_dir: str = "out",
                        recipient: str | None = None) -> dict:
    """Render the plan email (+ XLSX), send it if SMTP is configured, and record it."""
    subject = (f"{'UPDATED ' if plan['version'] > 1 else ''}Dispatch Plan {plan['plan_date']} "
               f"v{plan['version']} | {len(plan['assignments'])} loads, "
               f"{len(plan.get('holds', []))} on hold")
    html = build_email_html(plan, meta)
    to_list = resolve_recipients(meta, recipient)

    os.makedirs(out_dir, exist_ok=True)
    xlsx_path = build_xlsx(plan, os.path.join(out_dir, f"dispatch_plan_{plan['id']}.xlsx"))

    delivered, error = False, None
    if to_list and smtp_configured():
        user = os.environ["SMTP_USER"]
        try:
            msg = MIMEMultipart()
            msg["Subject"] = subject
            msg["From"] = user
            msg["To"] = ", ".join(to_list)
            msg.attach(MIMEText(html, "html"))
            if xlsx_path:
                with open(xlsx_path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition",
                                f"attachment; filename={os.path.basename(xlsx_path)}")
                msg.attach(part)
            _deliver(msg, user)
            delivered = True
        except Exception as exc:            # a mail problem must never kill a run
            error = f"{type(exc).__name__}: {exc}"[:200]
            print(f"[notify] SMTP send failed: {error}")
    elif not to_list:
        error = "no recipient set"
    else:
        error = "SMTP not configured"

    record = {
        "id": f"email-{plan['id']}",
        "plan_id": plan["id"],
        "subject": subject,
        "html": html,
        "attachment": os.path.basename(xlsx_path) if xlsx_path else None,
        "to": ", ".join(to_list) if to_list else "(no recipient set)",
        "recipients": to_list,
        "delivered": delivered,
        "error": error,
        "created_at": plan["generated_at"],
    }
    store.upsert("emails", record)
    return record
